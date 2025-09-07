"""
Fixed Flask canteen app with modern Bootstrap GUI + factuality page.

How it works:
- When you run this script, it will create a `templates/` and `static/` folder (if missing)
  and write all HTML templates and a simple CSS file into them.
- The Flask routes remain similar to your original app but with bug fixes:
  * order ID generation uses max(existing_id)+1
  * safer file reads/writes with exceptions handled
  * session role checks centralized
  * improved flash message display and input sanitization
- A new `/factuality` page (route: `/factuality`) is added with a modern card UI where
  admins can post short notes about facts/updates; it's editable by admin only.

Run: python canteen_app_fixed_and_ui.py
Open: http://127.0.0.1:5000

Note: This single-file script creates all templates for you. Do NOT manually copy-paste templates unless you want to customize them.
"""

import os
import json
from datetime import datetime
from flask import Flask, render_template, request, redirect, url_for, flash, session, abort, send_file, make_response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO

app = Flask(__name__)
app.secret_key = os.environ.get('FLASK_SECRET', 'change_this_secret')
import smtplib
from email.mime.text import MIMEText
from flask import url_for

# Your Gmail
EMAIL_ADDRESS = "harishgowda1379@gmail.com"
EMAIL_PASSWORD = "jewb jmhc igcq zjyv"  # not your Gmail password, use App Password
ADMIN_EMAIL = "kannadagamer387@gmail.com"

def send_email(subject, recipient, body):
    msg = MIMEText(body, "html")
    msg["Subject"] = subject
    msg["From"] = EMAIL_ADDRESS
    msg["To"] = recipient

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
        smtp.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
        smtp.send_message(msg)


# Data files
DATA_DIR = 'data'
ORDERS_FILE = os.path.join(DATA_DIR, 'orders.json')
USERS_FILE = os.path.join(DATA_DIR, 'users.json')
FACTUALITY_FILE = os.path.join(DATA_DIR, 'factuality.json')

# Template and static generation (writes modern Bootstrap templates on first run)
TEMPLATES = {
    'layout.html': '''<!doctype html>
<html lang="en">
  <head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>{{ title if title else 'Canteen Orders' }}</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
    <link href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.css" rel="stylesheet">
    <style>
      body { padding-top: 56px; }
      .card-modern { border-radius: 0.9rem; box-shadow: 0 6px 18px rgba(0,0,0,0.08); }
      .nav-brand { font-weight: 700; letter-spacing: .4px; }
      footer { padding: 1.2rem 0; }
    </style>
  </head>
  <body>
    <nav class="navbar navbar-expand-lg navbar-dark bg-primary fixed-top">
      <div class="container">
        <a class="navbar-brand nav-brand" href="{{ url_for('home') }}">CanteenHub</a>
        <button class="navbar-toggler" type="button" data-bs-toggle="collapse" data-bs-target="#navMenu">
          <span class="navbar-toggler-icon"></span>
        </button>

        <div class="collapse navbar-collapse" id="navMenu">
          <ul class="navbar-nav ms-auto">
            {% if session.get('username') %}
            <li class="nav-item"><a class="nav-link" href="#">{{ session.get('username') }}</a></li>
            <li class="nav-item"><a class="nav-link" href="{{ url_for('change_password') }}">Change Password</a></li>
            <li class="nav-item"><a class="nav-link" href="{{ url_for('logout') }}">Logout</a></li>
            {% else %}
            <li class="nav-item"><a class="nav-link" href="{{ url_for('home') }}">Login</a></li>
            {% endif %}
            <li class="nav-item"><a class="nav-link" href="{{ url_for('factuality') }}">Factuality</a></li>
          </ul>
        </div>
      </div>
    </nav>

    <main class="container mt-4">
      {% with messages = get_flashed_messages(with_categories=true) %}
        {% if messages %}
          <div class="mt-2">
            {% for category, msg in messages %}
              <div class="alert alert-{{ 'success' if category=='success' else ('danger' if category=='error' else 'info') }} alert-dismissible fade show" role="alert">
                {{ msg }}
                <button type="button" class="btn-close" data-bs-dismiss="alert"></button>
              </div>
            {% endfor %}
          </div>
        {% endif %}
      {% endwith %}

      {% block content %}{% endblock %}
    </main>

    <footer class="text-center mt-5">
      <div class="container small text-muted">© {{ now.year }} St. Francis College Canteen Management</div>
    </footer>

    <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/js/bootstrap.bundle.min.js"></script>
  </body>
</html>
''',

    'login_select.html': '''{% extends 'layout.html' %}
{% block content %}
<div class="row justify-content-center">
  <div class="col-md-8">
    <div class="card card-modern p-4">
      <h3 class="mb-3">Sign in as</h3>
      <div class="d-flex gap-3 flex-wrap">
        <a class="btn btn-outline-primary" href="{{ url_for('faculty_login') }}">Faculty</a>
        <a class="btn btn-outline-secondary" href="{{ url_for('admin_login') }}">Admin</a>
        <a class="btn btn-outline-success" href="{{ url_for('canteen_login') }}">Canteen</a>
      </div>
    </div>
  </div>
</div>
{% endblock %}
''',

    'faculty_login.html': '''{% extends 'layout.html' %}
{% block content %}
<div class="row justify-content-center">
  <div class="col-md-6">
    <div class="card card-modern p-4">
      <h4>Faculty Login</h4>
      <form method="post">
        <div class="mb-3"><label class="form-label">Username</label><input class="form-control" name="username" required></div>
        <div class="mb-3"><label class="form-label">Password</label><input class="form-control" name="password" type="password" required></div>
        <button class="btn btn-primary">Sign in</button>
      </form>
    </div>
  </div>
</div>
{% endblock %}
''',

    'admin_login.html': '''{% extends 'layout.html' %}
{% block content %}
<div class="row justify-content-center">
  <div class="col-md-6">
    <div class="card card-modern p-4">
      <h4>Admin Login</h4>
      <form method="post">
        <div class="mb-3"><label class="form-label">Username</label><input class="form-control" name="username" required></div>
        <div class="mb-3"><label class="form-label">Password</label><input class="form-control" name="password" type="password" required></div>
        <button class="btn btn-primary">Sign in</button>
      </form>
    </div>
  </div>
</div>
{% endblock %}
''',

    'canteen_login.html': '''{% extends 'layout.html' %}
{% block content %}
<div class="row justify-content-center">
  <div class="col-md-6">
    <div class="card card-modern p-4">
      <h4>Canteen Login</h4>
      <form method="post">
        <div class="mb-3"><label class="form-label">Username</label><input class="form-control" name="username" required></div>
        <div class="mb-3"><label class="form-label">Password</label><input class="form-control" name="password" type="password" required></div>
        <button class="btn btn-primary">Sign in</button>
      </form>
    </div>
  </div>
</div>
{% endblock %}
''',

    'order.html': '''{% extends 'layout.html' %}
{% block content %}
<div class="row">
  <div class="col-md-8">
    <div class="card card-modern p-4 mb-4">
      <h4>Place an Order</h4>
      <form method="post">
        <div class="row">
          <div class="col-md-6 mb-3"><label class="form-label">Department</label><input class="form-control" name="department"></div>
          <div class="col-md-6 mb-3"><label class="form-label">Event name</label><input class="form-control" name="eventName"></div>
        </div>
        <div class="row">
          <div class="col-md-4 mb-3"><label class="form-label">Date</label><input class="form-control" type="date" name="eventDate"></div>
          <div class="col-md-4 mb-3"><label class="form-label">Time</label><input class="form-control" type="time" name="eventTime"></div>
          <div class="col-md-4 mb-3"><label class="form-label">No. of persons</label><input class="form-control" type="number" name="numberOfPersons"></div>
        </div>
        <div class="mb-3"><label class="form-label">Items required (comma separated)</label><input class="form-control" name="itemsRequired"></div>
        <div class="row">
          <div class="col-md-6 mb-3"><label class="form-label">Cost per unit</label><input class="form-control" name="costPerUnit" type="number" step="0.01"></div>
          <div class="col-md-6 mb-3"><label class="form-label">Estimated amount</label><input class="form-control" name="estimatedAmount" type="number" step="0.01"></div>
        </div>
        <div class="mb-3"><label class="form-label">Remarks</label><textarea class="form-control" name="remarks"></textarea></div>
        <button class="btn btn-success">Submit order</button>
      </form>
    </div>

    <div class="card card-modern p-3">
      <h5>Your recent orders</h5>
      {% if orders %}
        <div class="list-group">
          {% for o in orders|reverse %}
            <div class="list-group-item">
              <div class="d-flex justify-content-between">
                <div>
                  <strong>#{{ o.id }} {{ o.event_name or '—' }}</strong> <small class="text-muted">({{ o.status }})</small>
                  <div class="small text-muted">{{ o.timestamp }}</div>
                </div>
                <div>
                  <small>By: {{ o.ordered_by }}</small>
                </div>
              </div>
            </div>
          {% endfor %}
        </div>
      {% else %}
        <div class="text-muted">No orders yet.</div>
      {% endif %}
    </div>
  </div>

  <div class="col-md-4">
    <div class="card card-modern p-3">
      <h6>Quick Tips</h6>
      <ul>
        <li>Use accurate headcounts to avoid food waste.</li>
        <li>Enter approximate costs for admin review.</li>
        <li>Admin will approve or reject orders.</li>
      </ul>
    </div>
  </div>
</div>
{% endblock %}
''',

    'admin_dashboard.html': '''{% extends 'layout.html' %}
{% block content %}
<!-- Dashboard Header -->
<div class="row mb-5">
    <div class="col-12">
        <div class="glass-card p-4 text-center animate-fadeIn">
            <h2 class="gradient-text mb-2">
                <i class="fas fa-tachometer-alt me-3"></i>
                Admin Dashboard
            </h2>
            <p class="text-muted mb-0">Manage and monitor all canteen operations</p>
        </div>
    </div>
</div>

<!-- Statistics Cards -->
<div class="row g-4 mb-5">
    <div class="col-lg-3 col-md-6">
        <div class="stat-card animate-fadeIn delay-1">
            <div class="d-flex align-items-center justify-content-between">
                <div>
                    <div class="stat-number">{{ orders|length }}</div>
                    <div class="stat-label">Total Orders</div>
                </div>
                <div class="opacity-50">
                    <i class="fas fa-clipboard-list" style="font-size: 2rem;"></i>
                </div>
            </div>
        </div>
    </div>
    
    <div class="col-lg-3 col-md-6">
        <div class="stat-card animate-fadeIn delay-2">
            <div class="d-flex align-items-center justify-content-between">
                <div>
                    <div class="stat-number">{{ orders|selectattr('status', 'equalto', 'pending')|list|length }}</div>
                    <div class="stat-label">Pending Review</div>
                </div>
                <div class="opacity-50">
                    <i class="fas fa-clock" style="font-size: 2rem;"></i>
                </div>
            </div>
        </div>
    </div>
    
    <div class="col-lg-3 col-md-6">
        <div class="stat-card animate-fadeIn delay-3">
            <div class="d-flex align-items-center justify-content-between">
                <div>
                    <div class="stat-number">{{ orders|selectattr('status', 'equalto', 'approved')|list|length }}</div>
                    <div class="stat-label">Approved</div>
                </div>
                <div class="opacity-50">
                    <i class="fas fa-check-circle" style="font-size: 2rem;"></i>
                </div>
            </div>
        </div>
    </div>
    
    <div class="col-lg-3 col-md-6">
        <div class="stat-card animate-fadeIn delay-4">
            <div class="d-flex align-items-center justify-content-between">
                <div>
                    <div class="stat-number">{{ ((orders|selectattr('status', 'equalto', 'approved')|list|length / orders|length * 100) if orders|length > 0 else 0)|round|int }}%</div>
                    <div class="stat-label">Success Rate</div>
                </div>
                <div class="opacity-50">
                    <i class="fas fa-chart-line" style="font-size: 2rem;"></i>
                </div>
            </div>
        </div>
    </div>
</div>

<!-- Quick Actions -->
<div class="row mb-4">
    <div class="col-12">
        <div class="glass-card p-4 animate-fadeIn delay-5">
            <div class="d-flex flex-wrap align-items-center justify-content-between">
                <h5 class="text-gradient mb-0">
                    <i class="fas fa-bolt me-2"></i>
                    Quick Actions
                </h5>
                <div class="d-flex flex-wrap gap-2 mt-3 mt-md-0">
                    <button class="btn btn-primary-modern btn-modern" onclick="approveAllPending()" data-tooltip="Approve all pending orders">
                        <i class="fas fa-check-double me-2"></i>
                        Approve All Pending
                    </button>
                    <a href="{{ url_for('export_orders_excel') }}" class="btn btn-success-modern btn-modern" data-tooltip="Export orders to Excel">
                        <i class="fas fa-file-excel me-2"></i>
                        Export to Excel
                    </a>
                    <button class="btn btn-warning-modern btn-modern" onclick="refreshDashboard()" data-tooltip="Refresh dashboard data">
                        <i class="fas fa-sync-alt me-2"></i>
                        Refresh
                    </button>
                </div>
            </div>
        </div>
    </div>
</div>

<!-- Orders Management Section -->
<div class="row">
    <div class="col-12">
        <div class="glass-card p-4 animate-fadeIn delay-6">
            <div class="d-flex align-items-center justify-content-between mb-4">
                <h4 class="text-gradient mb-0">
                    <i class="fas fa-list-alt me-2"></i>
                    Orders Management
                </h4>
                <div class="d-flex align-items-center gap-3">
                    <!-- Search Filter -->
                    <div class="input-group" style="width: 300px;">
                        <input type="text" class="form-control-modern" id="searchFilter" 
                               placeholder="Search orders..." style="border-radius: 12px 0 0 12px;">
                        <button class="btn btn-primary-modern" type="button" style="border-radius: 0 12px 12px 0;">
                            <i class="fas fa-search"></i>
                        </button>
                    </div>
                </div>
            </div>

            <!-- Modern Tab Navigation -->
            <div class="d-flex flex-wrap gap-2 mb-4">
                <button class="btn btn-modern btn-primary-modern active" data-filter="all" onclick="filterOrders('all')">
                    <i class="fas fa-list me-2"></i>All Orders
                    <span class="badge bg-light text-dark ms-2">{{ orders|length }}</span>
                </button>
                <button class="btn btn-modern btn-outline-warning" data-filter="pending" onclick="filterOrders('pending')">
                    <i class="fas fa-clock me-2"></i>Pending
                    <span class="badge bg-warning ms-2">{{ orders|selectattr('status', 'equalto', 'pending')|list|length }}</span>
                </button>
                <button class="btn btn-modern btn-outline-success" data-filter="approved" onclick="filterOrders('approved')">
                    <i class="fas fa-check me-2"></i>Approved
                    <span class="badge bg-success ms-2">{{ orders|selectattr('status', 'equalto', 'approved')|list|length }}</span>
                </button>
                <button class="btn btn-modern btn-outline-danger" data-filter="rejected" onclick="filterOrders('rejected')">
                    <i class="fas fa-times me-2"></i>Rejected
                    <span class="badge bg-danger ms-2">{{ orders|selectattr('status', 'equalto', 'rejected')|list|length }}</span>
                </button>
            </div>

            <!-- Orders Table -->
            <div class="table-responsive">
                {% if orders %}
                <table class="table-modern table" id="ordersTable">
                    <thead>
                        <tr>
                            <th>
                                <i class="fas fa-calendar-alt me-2"></i>
                                Event Details
                            </th>
                            <th>
                                <i class="fas fa-building me-2"></i>
                                Department
                            </th>
                            <th>
                                <i class="fas fa-user me-2"></i>
                                Requested By
                            </th>
                            <th>
                                <i class="fas fa-info-circle me-2"></i>
                                Status
                            </th>
                            <th>
                                <i class="fas fa-rupee-sign me-2"></i>
                                Amount
                            </th>
                            {% if session.get('role') == 'admin' %}
                            <th>
                                <i class="fas fa-cogs me-2"></i>
                                Actions
                            </th>
                            {% endif %}
                        </tr>
                    </thead>
                    <tbody>
                        {% for order in orders %}
                        <tr class="order-row animate-fadeIn" data-status="{{ order.status }}" style="animation-delay: {{ loop.index * 0.05 }}s;">
                            <td>
                                <div>
                                    <strong class="d-block">{{ order.event_name }}</strong>
                                    <small class="text-muted">
                                        <i class="fas fa-calendar me-1"></i>
                                        {{ order.event_date }}
                                        {% if order.event_time %}
                                        <br><i class="fas fa-clock me-1"></i>{{ order.event_time }}
                                        {% endif %}
                                    </small>
                                </div>
                            </td>
                            <td>
                                <span class="fw-500">{{ order.department }}</span>
                            </td>
                            <td>
                                <div class="d-flex align-items-center">
                                    <div class="bg-primary rounded-circle d-flex align-items-center justify-content-center me-2" 
                                         style="width: 32px; height: 32px; font-size: 0.8rem; color: white;">
                                        {{ order.ordered_by[0]|upper }}
                                    </div>
                                    <span class="fw-500">{{ order.ordered_by }}</span>
                                </div>
                            </td>
                            <td>
                                <span class="status-badge status-{{ order.status }}">
                                    <i class="fas fa-{{ 'clock' if order.status == 'pending' else 'check' if order.status == 'approved' else 'times' }} me-1"></i>
                                    {{ order.status|capitalize }}
                                </span>
                            </td>
                            <td>
                                <strong class="text-success">
                                    {% if order.estimated_amount %}
                                    ₹{{ "{:,.0f}".format(order.estimated_amount) }}
                                    {% else %}
                                    <span class="text-muted">Not specified</span>
                                    {% endif %}
                                </strong>
                            </td>
                            {% if session.get('role') == 'admin' %}
                            <td>
                                <div class="d-flex gap-2">
                                    {% if order.status == 'pending' %}
                                    <button class="btn btn-success btn-modern btn-sm" 
                                            onclick="approveOrder({{ order.id }})" 
                                            data-tooltip="Approve this order">
                                        <i class="fas fa-check"></i>
                                    </button>
                                    <button class="btn btn-danger btn-modern btn-sm" 
                                            onclick="rejectOrder({{ order.id }})" 
                                            data-tooltip="Reject this order">
                                        <i class="fas fa-times"></i>
                                    </button>
                                    {% endif %}
                                    <button class="btn btn-primary btn-modern btn-sm" 
                                            onclick="viewOrderDetails({{ order.id }})" 
                                            data-tooltip="View order details">
                                        <i class="fas fa-eye"></i>
                                    </button>
                                </div>
                            </td>
                            {% endif %}
                        </tr>
                        {% endfor %}
                    </tbody>
                </table>
                {% else %}
                <div class="text-center py-5">
                    <div class="glass-card p-5 d-inline-block">
                        <i class="fas fa-inbox text-muted mb-3" style="font-size: 3rem;"></i>
                        <h5 class="text-muted">No Orders Found</h5>
                        <p class="text-muted mb-0">No orders have been submitted yet.</p>
                    </div>
                </div>
                {% endif %}
            </div>
        </div>
    </div>
</div>

<script>
// Dashboard functionality
let currentFilter = 'all';

function filterOrders(status) {
    currentFilter = status;
    
    // Update active button
    document.querySelectorAll('[data-filter]').forEach(btn => {
        btn.classList.remove('btn-primary-modern', 'active');
        btn.classList.add('btn-outline-secondary');
    });
    
    const activeBtn = document.querySelector(`[data-filter="${status}"]`);
    activeBtn.classList.remove('btn-outline-secondary');
    activeBtn.classList.add('btn-primary-modern', 'active');
    
    // Filter table rows
    const rows = document.querySelectorAll('.order-row');
    rows.forEach((row, index) => {
        const rowStatus = row.dataset.status;
        const shouldShow = status === 'all' || rowStatus === status;
        
        if (shouldShow) {
            row.style.display = '';
            row.style.animationDelay = `${index * 0.05}s`;
            row.classList.add('animate-fadeIn');
        } else {
            row.style.display = 'none';
        }
    });
    
    // Update empty state
    updateEmptyState(status);
}

function updateEmptyState(status) {
    const visibleRows = document.querySelectorAll('.order-row[style=""], .order-row:not([style])');
    const filteredRows = Array.from(visibleRows).filter(row => {
        return status === 'all' || row.dataset.status === status;
    });
    
    const table = document.getElementById('ordersTable');
    const emptyState = document.getElementById('emptyState');
    
    if (filteredRows.length === 0 && table) {
        if (!emptyState) {
            const emptyDiv = document.createElement('div');
            emptyDiv.id = 'emptyState';
            emptyDiv.className = 'text-center py-5';
            emptyDiv.innerHTML = `
                <div class="glass-card p-5 d-inline-block animate-fadeIn">
                    <i class="fas fa-search text-muted mb-3" style="font-size: 3rem;"></i>
                    <h5 class="text-muted">No ${status === 'all' ? '' : status.charAt(0).toUpperCase() + status.slice(1)} Orders Found</h5>
                    <p class="text-muted mb-0">Try adjusting your search or filters.</p>
                </div>
            `;
            table.parentNode.appendChild(emptyDiv);
        }
        table.style.display = 'none';
    } else {
        if (emptyState) {
            emptyState.remove();
        }
        if (table) {
            table.style.display = '';
        }
    }
}

// Search functionality
document.getElementById('searchFilter')?.addEventListener('input', function() {
    const searchTerm = this.value.toLowerCase();
    const rows = document.querySelectorAll('.order-row');
    
    rows.forEach(row => {
        const text = row.textContent.toLowerCase();
        const matchesSearch = text.includes(searchTerm);
        const matchesFilter = currentFilter === 'all' || row.dataset.status === currentFilter;
        
        row.style.display = (matchesSearch && matchesFilter) ? '' : 'none';
    });
});

// Admin actions
function approveOrder(orderId) {
    if (confirm('Are you sure you want to approve this order?')) {
        window.location.href = `/approve_order/${orderId}`;
    }
}

function rejectOrder(orderId) {
    const reason = prompt('Please provide a reason for rejection (optional):');
    if (confirm('Are you sure you want to reject this order?')) {
        // You can modify this to pass the reason if needed
        window.location.href = `/reject_order/${orderId}`;
    }
}

function approveAllPending() {
    const pendingCount = document.querySelectorAll('[data-status="pending"]').length;
    
    if (pendingCount === 0) {
        window.modernCanteenApp?.showNotification('No pending orders to approve', 'info');
        return;
    }
    
    if (confirm(`Are you sure you want to approve all ${pendingCount} pending orders?`)) {
        // This would need backend implementation
        window.modernCanteenApp?.showNotification('Feature coming soon! Please approve orders individually for now.', 'info');
    }
}

function viewOrderDetails(orderId) {
    // Create modal or navigate to details page
    window.modernCanteenApp?.showNotification('Order details view coming soon!', 'info');
}

function exportData() {
    // Excel export
    window.location.href = '/export_orders_excel';
    window.modernCanteenApp?.showNotification('Downloading Excel file...', 'success');
}

function refreshDashboard() {
    window.location.reload();
}

// Initialize dashboard
document.addEventListener('DOMContentLoaded', function() {
    // Add smooth animations to table rows
    const rows = document.querySelectorAll('.order-row');
    rows.forEach((row, index) => {
        row.style.animationDelay = `${index * 0.1}s`;
    });
    
    // Initialize filter
    filterOrders('all');
});
</script>

<style>
/* Dashboard-specific styles */
.btn-modern.btn-outline-warning {
    border-color: var(--warning-color);
    color: var(--warning-color);
    background: transparent;
}

.btn-modern.btn-outline-warning:hover {
    background: var(--warning-color);
    color: white;
}

.btn-modern.btn-outline-success {
    border-color: var(--accent-color);
    color: var(--accent-color);
    background: transparent;
}

.btn-modern.btn-outline-success:hover {
    background: var(--accent-color);
    color: white;
}

.btn-modern.btn-outline-danger {
    border-color: var(--danger-color);
    color: var(--danger-color);
    background: transparent;
}

.btn-modern.btn-outline-danger:hover {
    background: var(--danger-color);
    color: white;
}

.btn-modern.btn-outline-secondary {
    border-color: var(--border-color);
    color: var(--text-secondary);
    background: transparent;
}

.btn-modern.btn-outline-secondary:hover {
    background: var(--secondary-color);
    color: var(--primary-color);
    border-color: var(--primary-color);
}

/* Table enhancements */
.table-modern tbody tr {
    cursor: pointer;
}

.table-modern tbody tr:hover {
    transform: scale(1.01);
    box-shadow: 0 5px 15px rgba(0,0,0,0.1);
}

.order-row {
    transition: all 0.3s ease;
}

/* Filter button animations */
[data-filter] {
    transition: all 0.3s ease;
    position: relative;
    overflow: hidden;
}

[data-filter]:not(.active):hover {
    transform: translateY(-1px);
}

/* Quick actions styling */
.btn-sm.btn-modern {
    padding: 0.4rem 0.8rem;
    font-size: 0.8rem;
}

/* Search input styling */
.input-group .form-control-modern:focus {
    box-shadow: 0 0 0 3px rgba(99, 102, 241, 0.1);
    border-color: var(--primary-color);
}
</style>
{% endblock %}
'''

    'canteen_orders.html' '''{% extends 'layout.html' %}
{% block content %}
<h3>Canteen — Approved Orders</h3>
{% if orders %}
  <div class="row">
    {% for o in orders %}
      <div class="col-md-6 mb-3">
        <div class="card card-modern p-3">
          <div class="d-flex justify-content-between">
            <div>
              <strong>#{{ o.id }} {{ o.event_name or '' }}</strong>
              <div class="small text-muted">{{ o.department }} — {{ o.timestamp }}</div>
            </div>
            <div>
              <a class="btn btn-sm btn-primary" href="{{ url_for('complete_order', order_id=o.id) }}">Mark Completed</a>
            </div>
          </div>
          <div class="mt-2 small">Items: {{ o.items_required }}</div>
        </div>
      </div>
    {% endfor %}
  </div>
{% else %}
  <div class="text-muted">No approved orders yet.</div>
{% endif %}
{% endblock %}
''',

    'change_password.html': '''{% extends 'layout.html' %}
{% block content %}
<div class="row justify-content-center">
  <div class="col-md-6">
    <div class="card card-modern p-4">
      <h5>Change Password</h5>
      <form method="post">
        <div class="mb-3"><label class="form-label">Current password</label><input class="form-control" name="current_password" type="password" required></div>
        <div class="mb-3"><label class="form-label">New password</label><input class="form-control" name="new_password" type="password" required></div>
        <div class="mb-3"><label class="form-label">Confirm new password</label><input class="form-control" name="confirm_password" type="password" required></div>
        <button class="btn btn-primary">Change</button>
      </form>
    </div>
  </div>
</div>
{% endblock %}
''',

    'change_password_login.html': '''{% extends 'layout.html' %}
{% block content %}
<div class="row justify-content-center">
  <div class="col-md-6">
    <div class="card card-modern p-4">
      <h5>Reset Password (Login)</h5>
      <form method="post">
        <div class="mb-3"><label class="form-label">Username</label><input class="form-control" name="username" required></div>
        <div class="mb-3"><label class="form-label">Current password</label><input class="form-control" name="current_password" type="password" required></div>
        <div class="mb-3"><label class="form-label">New password</label><input class="form-control" name="new_password" type="password" required></div>
        <div class="mb-3"><label class="form-label">Confirm</label><input class="form-control" name="confirm_password" type="password" required></div>
        <button class="btn btn-primary">Reset</button>
      </form>
    </div>
  </div>
</div>
{% endblock %}
''',

    'factuality.html': '''{% extends 'layout.html' %}
{% block content %}
<div class="row">
  <div class="col-md-8">
    <div class="card card-modern p-3 mb-3">
      <h4 class="text-gradient">
        <i class="fas fa-bullhorn me-2"></i>
        Factuality / Notices
      </h4>
      <p class="text-muted">Important updates and announcements for all staff members.</p>
      {% if notices %}
        <div class="list-group list-group-flush">
          {% for n in notices|reverse %}
            <div class="list-group-item border-0 bg-light rounded mb-2 p-3">
              <div class="d-flex justify-content-between align-items-start">
                <div class="flex-grow-1">
                  <h6 class="fw-bold text-primary mb-1">
                    <i class="fas fa-info-circle me-2"></i>
                    {{ n.title }}
                  </h6>
                  <div class="small text-muted mb-2">
                    <i class="fas fa-clock me-1"></i>
                    {{ n.timestamp }}
                    {% if n.posted_by %}
                      • Posted by {{ n.posted_by }}
                    {% endif %}
                  </div>
                  <div class="mt-2">{{ n.body }}</div>
                  
                  <!-- Faculty can react to notices -->
                  {% if session.get('role') == 'faculty' %}
                  <div class="mt-3">
                    <button class="btn btn-sm btn-outline-primary me-2" onclick="reactToNotice('{{ loop.index0 }}', 'helpful')">
                      <i class="fas fa-thumbs-up me-1"></i>
                      Helpful
                    </button>
                    <button class="btn btn-sm btn-outline-secondary" onclick="reactToNotice('{{ loop.index0 }}', 'noted')">
                      <i class="fas fa-check me-1"></i>
                      Noted
                    </button>
                  </div>
                  {% endif %}
                </div>
                {% if session.get('role') == 'admin' %}
                  <div>
                    <a class="btn btn-sm btn-outline-danger" href="{{ url_for('delete_notice', index=loop.index0) }}">
                      <i class="fas fa-trash me-1"></i>
                      Delete
                    </a>
                  </div>
                {% endif %}
              </div>
            </div>
          {% endfor %}
        </div>
      {% else %}
        <div class="text-center text-muted py-4">
          <i class="fas fa-inbox fa-3x mb-3 opacity-50"></i>
          <p class="mb-0">No notices published yet.</p>
          <small>Check back later for important updates!</small>
        </div>
      {% endif %}
    </div>

    {% if session.get('role') == 'admin' %}
    <div class="card card-modern p-3">
      <h6 class="text-gradient">
        <i class="fas fa-plus-circle me-2"></i>
        Post New Notice
      </h6>
      <form method="post" action="{{ url_for('post_notice') }}" class="row g-3">
        <div class="col-12">
          <label class="form-label fw-bold">Notice Title</label>
          <input class="form-control form-control-lg" name="title" placeholder="e.g., Menu Update, Timing Change" required>
        </div>
        <div class="col-12">
          <label class="form-label fw-bold">Notice Content</label>
          <textarea class="form-control" name="body" placeholder="Provide detailed information about the update..." rows="4" required></textarea>
        </div>
        <div class="col-12">
          <button class="btn btn-primary btn-lg">
            <i class="fas fa-paper-plane me-2"></i>
            Publish Notice
          </button>
        </div>
      </form>
    </div>
    {% endif %}
  </div>
  
  <div class="col-md-4">
    <div class="card card-modern p-3 mb-3">
      <h6 class="text-gradient">
        <i class="fas fa-info-circle me-2"></i>
        About Notices
      </h6>
      <p class="small text-muted mb-3">This page displays important updates such as:</p>
      <ul class="small text-muted">
        <li>Menu changes and new items</li>
        <li>Canteen operating hours</li>
        <li>Special event announcements</li>
        <li>Holiday schedules</li>
        <li>Emergency notices</li>
      </ul>
    </div>
    
    {% if session.get('role') == 'faculty' %}
    <div class="card card-modern p-3">
      <h6 class="text-gradient">
        <i class="fas fa-user-graduate me-2"></i>
        Faculty Actions
      </h6>
      <p class="small text-muted mb-3">As faculty member, you can:</p>
      <ul class="small">
        <li class="text-success"><i class="fas fa-check me-1"></i> View all notices</li>
        <li class="text-success"><i class="fas fa-check me-1"></i> React to notices</li>
        <li class="text-primary"><i class="fas fa-plus me-1"></i> <a href="{{ url_for('order') }}" class="text-decoration-none">Place orders</a></li>
      </ul>
    </div>
    {% endif %}
  </div>
</div>

<script>
function reactToNotice(noticeIndex, reaction) {
    // For now, just show a notification
    // In a full implementation, this would save the reaction to the database
    if (window.modernCanteenApp) {
        window.modernCanteenApp.showNotification(
            `Thank you for marking this notice as "${reaction}"!`, 
            'success'
        );
    } else {
        alert(`Thank you for marking this notice as "${reaction}"!`);
    }
}
</script>
{% endblock %}
'''
}


def ensure_templates_written():
    if not os.path.isdir('templates'):
        os.makedirs('templates', exist_ok=True)
    for name, content in TEMPLATES.items():
        path = os.path.join('templates', name)
        if not os.path.exists(path):
            with open(path, 'w', encoding='utf-8') as f:
                f.write(content)

# Data helpers

def ensure_data_dir():
    os.makedirs(DATA_DIR, exist_ok=True)
    if not os.path.exists(ORDERS_FILE):
        with open(ORDERS_FILE, 'w') as f:
            json.dump([], f)
    if not os.path.exists(USERS_FILE):
        default_users = {
            "faculty": {"password": "faculty123", "role": "faculty"},
            "admin": {"password": "admin123", "role": "admin"},
            "canteen": {"password": "canteen123", "role": "canteen"}
        }
        with open(USERS_FILE, 'w') as f:
            json.dump(default_users, f, indent=2)
    if not os.path.exists(FACTUALITY_FILE):
        with open(FACTUALITY_FILE, 'w') as f:
            json.dump([], f)


def load_json(path):
    try:
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return []


def save_json(path, payload):
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)

# initialize on import
ensure_templates_written()
ensure_data_dir()

# Utility: role required decorator-like check

def require_role(role):
    def wrapper():
        if session.get('role') != role:
            flash('Access denied.', 'error')
            return redirect(url_for('home'))
        return None
    return wrapper

@app.context_processor
def inject_now():
    return {'now': datetime.now()}

# Routes
@app.route('/')
def home():
    return render_template('login_select.html')
import requests


@app.route('/faculty_login', methods=['GET', 'POST'])
def faculty_login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        users = load_json(USERS_FILE)
        user = users.get(username) if isinstance(users, dict) else None
        if user and user.get('password') == password and user.get('role') == 'faculty':
            session['username'] = username
            session['role'] = 'faculty'
            return redirect(url_for('order'))
        flash('Invalid credentials!', 'error')
    return render_template('faculty_login.html')


@app.route('/admin_login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        users = load_json(USERS_FILE)
        user = users.get(username) if isinstance(users, dict) else None
        if user and user.get('password') == password and user.get('role') == 'admin':
            session['username'] = username
            session['role'] = 'admin'
            return redirect(url_for('admin_dashboard'))
        flash('Invalid credentials!', 'error')
    return render_template('admin_login.html')


@app.route('/canteen_login', methods=['GET', 'POST'])
def canteen_login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        users = load_json(USERS_FILE)
        user = users.get(username) if isinstance(users, dict) else None
        if user and user.get('password') == password and user.get('role') == 'canteen':
            session['username'] = username
            session['role'] = 'canteen'
            return redirect(url_for('canteen_orders'))
        flash('Invalid credentials!', 'error')
    return render_template('canteen_login.html')


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('home'))


@app.route('/order', methods=['GET', 'POST'])
def order():
    if session.get('role') != 'faculty':
        return redirect(url_for('home'))

    orders = load_json(ORDERS_FILE)

    if request.method == 'POST':
        try:
            # generate new id
            existing_ids = [o.get('id', 0) for o in orders]
            new_id = (max(existing_ids) + 1) if existing_ids else 1

            # Handle multiple purposes
            purposes = request.form.getlist('purpose')
            custom_purpose = request.form.get('customPurpose', '').strip()
            if 'Others' in purposes and custom_purpose:
                purposes.append(f"Others: {custom_purpose}")
                purposes.remove('Others')
            
            order_data = {
                'id': new_id,
                'department': request.form.get('department', '').strip(),
                'event_name': request.form.get('eventName', '').strip(),
                'event_date': request.form.get('eventDate', ''),
                'event_time': request.form.get('eventTime', ''),
                'purpose': ', '.join(purposes) if purposes else '',
                'service_type': request.form.get('serviceType', ''),
                'number_of_persons': request.form.get('numberOfPersons', ''),
                'items_required': request.form.get('itemsRequired', '').strip(),
                'cost_per_unit': float(request.form.get('costPerUnit', 0) or 0),
                'estimated_amount': float(request.form.get('estimatedAmount', 0) or 0),
                'remarks': request.form.get('remarks', '').strip(),
                'status': 'pending',
                'timestamp': datetime.now().isoformat(sep=' ', timespec='seconds'),
                'ordered_by': session.get('username', ''),
                'approved_at': None,
                'approved_by': None,
                'completed_at': None,
                'completed_by': None
            }

            # Save order
            orders.append(order_data)
            save_json(ORDERS_FILE, orders)

            # === 📧 Send Email Notification to Admin ===
            DOMAIN = "https://sfccanteen.onrender.com"
            approve_link = f"{DOMAIN}{url_for('approve_order', order_id=new_id)}?token=secure123"
            reject_link = f"{DOMAIN}{url_for('reject_order', order_id=new_id)}?token=secure123"



            email_body = f"""
            <h2>📢 New Canteen Order</h2>
            <p>
            <b>Department:</b> {order_data['department']}<br>
            <b>Event:</b> {order_data['event_name']}<br>
            <b>Date:</b> {order_data['event_date']} at {order_data['event_time']}<br>
            <b>Purpose:</b> {order_data['purpose']}<br>
            <b>Service Type:</b> {order_data['service_type']}<br>
            <b>No. of Persons:</b> {order_data['number_of_persons']}<br>
            <b>Items:</b> {order_data['items_required']}<br>
            <b>Estimate:</b> ₹{order_data['estimated_amount']}<br>
            <b>Remarks:</b> {order_data['remarks']}
            </p>
            <p>
            <a href="{approve_link}" style="padding:12px 24px;background:#16a34a;color:#fff;
            text-decoration:none;border-radius:6px;font-weight:bold;">✅ Approve</a>
            &nbsp;&nbsp;
            <a href="{reject_link}" style="padding:12px 24px;background:#dc2626;color:#fff;
            text-decoration:none;border-radius:6px;font-weight:bold;">❌ Reject</a>
            </p>
            """

            send_email("📢 New Order Request", ADMIN_EMAIL, email_body)

            flash('Order submitted successfully! It will be reviewed by administration.', 'success')
            return redirect(url_for('order'))

        except Exception as e:
            flash('Failed to submit order: {}'.format(e), 'error')

    # for faculty view show their orders only
    user_orders = [o for o in orders if o.get('ordered_by') == session.get('username')]
    return render_template('order.html', orders=user_orders)


@app.route('/admin_dashboard')
def admin_dashboard():
    if session.get('role') != 'admin':
        return redirect(url_for('home'))
    orders = load_json(ORDERS_FILE)
    return render_template('admin_dashboard.html', orders=orders)


@app.route('/approve_order/<int:order_id>')
def approve_order(order_id):
    token = request.args.get("token")
    if session.get('role') != 'admin' and token != "secure123":
        return "Access denied!", 403

    orders = load_json(ORDERS_FILE)
    order_found = False

    for order in orders:
        if int(order.get('id')) == order_id:
            order['status'] = 'approved'
            order['approved_at'] = datetime.now().isoformat(sep=' ', timespec='seconds')
            order['approved_by'] = session.get('username') if session.get('role') == 'admin' else "email-approval"
            order_found = True
            break

    if order_found:
        save_json(ORDERS_FILE, orders)
        if token == "secure123":
            return f"""
            <html>
            <body style="font-family:Arial,sans-serif;text-align:center;padding:40px;">
            <h2 style="color:#16a34a;">✅ Order Approved!</h2>
            <p>Order #{order_id} has been successfully approved.</p>
            </body>
            </html>
            """
        else:
            flash(f'Order #{order_id} approved successfully!', 'success')
            return redirect(url_for('admin_dashboard'))
    else:
        if token == "secure123":
            return f"⚠️ Order #{order_id} not found!"
        else:
            flash(f'Order #{order_id} not found!', 'error')
            return redirect(url_for('admin_dashboard'))


@app.route('/reject_order/<int:order_id>')
def reject_order(order_id):
    token = request.args.get("token")
    if session.get('role') != 'admin' and token != "secure123":
        return "Access denied!", 403

    orders = load_json(ORDERS_FILE)
    order_found = False

    for order in orders:
        if int(order.get('id')) == order_id:
            order['status'] = 'rejected'
            order['approved_at'] = datetime.now().isoformat(sep=' ', timespec='seconds')
            order['approved_by'] = session.get('username') if session.get('role') == 'admin' else "email-approval"
            order_found = True
            break

    if order_found:
        save_json(ORDERS_FILE, orders)
        if token == "secure123":
            return f"""
            <html>
            <body style="font-family:Arial,sans-serif;text-align:center;padding:40px;">
            <h2 style="color:#dc2626;">❌ Order Rejected!</h2>
            <p>Order #{order_id} has been rejected.</p>
            </body>
            </html>
            """
        else:
            flash(f'Order #{order_id} rejected successfully!', 'success')
            return redirect(url_for('admin_dashboard'))
    else:
        if token == "secure123":
            return f"⚠️ Order #{order_id} not found!"
        else:
            flash(f'Order #{order_id} not found!', 'error')
            return redirect(url_for('admin_dashboard'))


@app.route('/approve_all_pending')
def approve_all_pending():
    if session.get('role') != 'admin':
        flash('Access denied!', 'error')
        return redirect(url_for('home'))
    
    orders = load_json(ORDERS_FILE)
    approved_count = 0
    current_time = datetime.now().isoformat(sep=' ', timespec='seconds')
    current_user = session.get('username')
    
    for order in orders:
        if order.get('status') == 'pending':
            order['status'] = 'approved'
            order['approved_at'] = current_time
            order['approved_by'] = current_user
            approved_count += 1
    
    if approved_count > 0:
        save_json(ORDERS_FILE, orders)
        flash(f'Successfully approved {approved_count} pending order{"s" if approved_count > 1 else ""}! They have been sent to canteen for fulfillment.', 'success')
    else:
        flash('No pending orders found to approve.', 'info')
    
    return redirect(url_for('admin_dashboard'))


@app.route('/canteen_orders')
def canteen_orders():
    if session.get('role') != 'canteen':
        return redirect(url_for('home'))
    orders = load_json(ORDERS_FILE)
    approved_orders = [o for o in orders if o.get('status') == 'approved']
    return render_template('canteen_orders.html', orders=approved_orders)


@app.route('/complete_order/<int:order_id>')
def complete_order(order_id):
    if session.get('role') != 'canteen':
        flash('Access denied!', 'error')
        return redirect(url_for('home'))
    
    orders = load_json(ORDERS_FILE)
    order_found = False
    
    for order in orders:
        if int(order.get('id')) == order_id and order.get('status') == 'approved':
            order['status'] = 'completed'
            order['completed_at'] = datetime.now().isoformat(sep=' ', timespec='seconds')
            order['completed_by'] = session.get('username')
            order_found = True
            break
    
    if order_found:
        save_json(ORDERS_FILE, orders)
        flash(f'Order #{order_id} marked as completed successfully!', 'success')
    else:
        flash(f'Order #{order_id} not found or not approved!', 'error')
    
    return redirect(url_for('canteen_orders'))


@app.route('/complete_all_orders')
def complete_all_orders():
    if session.get('role') != 'canteen':
        flash('Access denied!', 'error')
        return redirect(url_for('home'))
    
    orders = load_json(ORDERS_FILE)
    completed_count = 0
    current_time = datetime.now().isoformat(sep=' ', timespec='seconds')
    current_user = session.get('username')
    
    for order in orders:
        if order.get('status') == 'approved':
            order['status'] = 'completed'
            order['completed_at'] = current_time
            order['completed_by'] = current_user
            completed_count += 1
    
    if completed_count > 0:
        save_json(ORDERS_FILE, orders)
        flash(f'Successfully completed {completed_count} order{"s" if completed_count > 1 else ""}!', 'success')
    else:
        flash('No approved orders found to complete.', 'info')
    
    return redirect(url_for('canteen_orders'))


@app.route('/change_password', methods=['GET', 'POST'])
def change_password():
    if 'role' not in session:
        return redirect(url_for('home'))
    if request.method == 'POST':
        current_password = request.form.get('current_password', '')
        new_password = request.form.get('new_password', '')
        confirm_password = request.form.get('confirm_password', '')
        users = load_json(USERS_FILE)
        username = session.get('username')
        user = users.get(username) if isinstance(users, dict) else None
        if user and user.get('password') == current_password:
            if new_password == confirm_password:
                user['password'] = new_password
                users[username] = user
                save_json(USERS_FILE, users)
                flash('Password changed successfully!', 'success')
                return redirect(url_for('change_password'))
            else:
                flash('New passwords do not match!', 'error')
        else:
            flash('Current password is incorrect!', 'error')
    return render_template('change_password.html')


@app.route('/change_password_login', methods=['GET', 'POST'])
def change_password_login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        current_password = request.form.get('current_password', '')
        new_password = request.form.get('new_password', '')
        confirm_password = request.form.get('confirm_password', '')
        users = load_json(USERS_FILE)
        user = users.get(username) if isinstance(users, dict) else None
        if user and user.get('password') == current_password:
            if new_password == confirm_password:
                user['password'] = new_password
                users[username] = user
                save_json(USERS_FILE, users)
                flash('Password changed successfully! You can now login with your new password.', 'success')
                return redirect(url_for('home'))
            else:
                flash('New passwords do not match!', 'error')
        else:
            flash('Username or current password is incorrect!', 'error')
    return render_template('change_password_login.html')


# Factuality / notices
@app.route('/factuality')
def factuality():
    notices = load_json(FACTUALITY_FILE)
    return render_template('factuality.html', notices=notices)


@app.route('/post_notice', methods=['POST'])
def post_notice():
    if session.get('role') != 'admin':
        flash('Only admins can post notices.', 'error')
        return redirect(url_for('factuality'))
    title = request.form.get('title', '').strip()
    body = request.form.get('body', '').strip()
    if not title or not body:
        flash('Title and body are required.', 'error')
        return redirect(url_for('factuality'))
    notices = load_json(FACTUALITY_FILE)
    notices.append({'title': title, 'body': body, 'timestamp': datetime.now().isoformat(sep=' ', timespec='seconds'), 'posted_by': session.get('username')})
    save_json(FACTUALITY_FILE, notices)
    flash('Notice posted.', 'success')
    return redirect(url_for('factuality'))


@app.route('/delete_notice/<int:index>')
def delete_notice(index):
    if session.get('role') != 'admin':
        flash('Only admins can delete notices.', 'error')
        return redirect(url_for('factuality'))
    notices = load_json(FACTUALITY_FILE)
    try:
        # index from template loop.index0 refers to reversed list; but kept simple: delete by index in stored order
        if 0 <= index < len(notices):
            notices.pop(index)
            save_json(FACTUALITY_FILE, notices)
            flash('Notice deleted.', 'success')
    except Exception:
        flash('Could not delete notice.', 'error')
    return redirect(url_for('factuality'))


@app.route('/export_orders_excel')
def export_orders_excel():
    """Export all orders to Excel file for admin"""
    if session.get('role') != 'admin':
        flash('Access denied!', 'error')
        return redirect(url_for('home'))
    
    orders = load_json(ORDERS_FILE)
    
    if not orders:
        flash('No orders to export!', 'warning')
        return redirect(url_for('admin_dashboard'))
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Canteen Orders"
    
    # Define headers
    headers = [
        'Order ID', 'Department', 'Event Name', 'Event Date', 'Event Time',
        'Purpose', 'Service Type', 'Number of Persons', 'Items Required',
        'Cost Per Unit', 'Estimated Amount', 'Remarks', 'Status',
        'Ordered By', 'Order Date', 'Approved By', 'Approved Date',
        'Completed By', 'Completed Date'
    ]
    
    # Style for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Add data rows
    for row, order in enumerate(orders, 2):
        ws.cell(row=row, column=1, value=order.get('id', ''))
        ws.cell(row=row, column=2, value=order.get('department', ''))
        ws.cell(row=row, column=3, value=order.get('event_name', ''))
        ws.cell(row=row, column=4, value=order.get('event_date', ''))
        ws.cell(row=row, column=5, value=order.get('event_time', ''))
        ws.cell(row=row, column=6, value=order.get('purpose', ''))
        ws.cell(row=row, column=7, value=order.get('service_type', ''))
        ws.cell(row=row, column=8, value=order.get('number_of_persons', ''))
        ws.cell(row=row, column=9, value=order.get('items_required', ''))
        ws.cell(row=row, column=10, value=order.get('cost_per_unit', ''))
        ws.cell(row=row, column=11, value=order.get('estimated_amount', ''))
        ws.cell(row=row, column=12, value=order.get('remarks', ''))
        ws.cell(row=row, column=13, value=order.get('status', ''))
        ws.cell(row=row, column=14, value=order.get('ordered_by', ''))
        ws.cell(row=row, column=15, value=order.get('timestamp', ''))
        ws.cell(row=row, column=16, value=order.get('approved_by', ''))
        ws.cell(row=row, column=17, value=order.get('approved_at', ''))
        ws.cell(row=row, column=18, value=order.get('completed_by', ''))
        ws.cell(row=row, column=19, value=order.get('completed_at', ''))
    
     
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create BytesIO object to store the file
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename with current date
    filename = f"canteen_orders_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    # Create response
    response = make_response(output.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    
    flash(f'Excel file "{filename}" downloaded successfully!', 'success')
    return response


# Department Management Routes
@app.route('/manage_departments')
def manage_departments():
    if session.get('role') != 'admin':
        flash('Access denied!', 'error')
        return redirect(url_for('home'))
    
    users = load_json(USERS_FILE)
    # Filter out system accounts (admin, canteen)
    departments = {k: v for k, v in users.items() if k not in ['admin', 'canteen'] and v.get('role') == 'faculty'}
    return render_template('manage_departments.html', departments=departments)


@app.route('/add_department', methods=['POST'])
def add_department():
    if session.get('role') != 'admin':
        flash('Access denied!', 'error')
        return redirect(url_for('home'))
    
    department_name = request.form.get('department_name', '').strip()
    password = request.form.get('password', '').strip()
    
    if not department_name or not password:
        flash('Department name and password are required!', 'error')
        return redirect(url_for('manage_departments'))
    
    users = load_json(USERS_FILE)
    
    # Check if department already exists
    if department_name.lower() in users:
        flash(f'Department "{department_name}" already exists!', 'error')
        return redirect(url_for('manage_departments'))
    
    # Add new department
    users[department_name.lower()] = {
        'password': password,
        'role': 'faculty',
        'department_name': department_name,
        'created_at': datetime.now().isoformat(sep=' ', timespec='seconds'),
        'created_by': session.get('username')
    }
    
    save_json(USERS_FILE, users)
    flash(f'Department "{department_name}" added successfully!', 'success')
    return redirect(url_for('manage_departments'))


@app.route('/delete_department/<department_id>')
def delete_department(department_id):
    if session.get('role') != 'admin':
        flash('Access denied!', 'error')
        return redirect(url_for('home'))
    
    users = load_json(USERS_FILE)
    
    if department_id in ['admin', 'canteen']:
        flash('Cannot delete system accounts!', 'error')
        return redirect(url_for('manage_departments'))
    
    if department_id in users:
        department_name = users[department_id].get('department_name', department_id)
        del users[department_id]
        save_json(USERS_FILE, users)
        flash(f'Department "{department_name}" deleted successfully!', 'success')
    else:
        flash('Department not found!', 'error')
    
    return redirect(url_for('manage_departments'))


@app.route('/reset_department_password/<department_id>', methods=['POST'])
def reset_department_password(department_id):
    if session.get('role') != 'admin':
        flash('Access denied!', 'error')
        return redirect(url_for('home'))
    
    new_password = request.form.get('new_password', '').strip()
    
    if not new_password:
        flash('New password is required!', 'error')
        return redirect(url_for('manage_departments'))
    
    users = load_json(USERS_FILE)
    
    if department_id in users:
        users[department_id]['password'] = new_password
        save_json(USERS_FILE, users)
        department_name = users[department_id].get('department_name', department_id)
        flash(f'Password for "{department_name}" updated successfully!', 'success')
    else:
        flash('Department not found!', 'error')
    
    return redirect(url_for('manage_departments'))


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))  # Use Render's PORT if available
    app.run(debug=True, host='0.0.0.0', port=port)
ṣ